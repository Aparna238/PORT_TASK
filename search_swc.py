import os
import json
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

new_data = []
str_data= os.environ.get('DATA7', '')
data7 = json.loads(str_data)

swc_name = os.environ.get('SWC_NAME','')

for row in data7:
    if swc_name == row['Swc_component']:
        new_data.append({'Swc_component':row['Swc_component'], 'Swc_port' : row['Swc_port'], 'Interface_name': row ['Interface_name'], 'interface_type': row ['interface_type'], 'Category': row['Category']})

# df7 = pd.DataFrame(new_data)
# df7.to_excel("swcname.xlsx")
df = pd.DataFrame(new_data)
wb = load_workbook("Port_Data.xlsx")
new_sheet = wb.create_sheet(title=swc_name)

# for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
# new_sheet.append(row)

data = dataframe_to_rows(df, index=False, header=True)
for rows in data:
    new_sheet.append(rows)

wb.save("Port_Data.xlsx")